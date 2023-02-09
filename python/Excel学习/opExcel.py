# openpyxl是一个非标准库，因此需要自行安装，安装过程并不困难，Windows/Mac用户均可以在命令行(CMD)/终端(Terminal)中使用pip安装

# pip install openpyxl

# 前置知识
# 单元格Cell
# 列column
# 行row
# 表sheet

# 1.载入Excel       注意load_workbook只能打开已经存在的Excel，不能创建新的工作簿
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Side, Border
from openpyxl.styles import PatternFill , GradientFill
fileName = 'C:/Users/xiaojiang/Desktop/大杂烩/gitP/MyNote/python/Excel学习/Excel学习.xlsx'
workbook = load_workbook(fileName)
print(workbook.sheetnames)
# 2.根据名称获取工作表
# 如果只有一张工作表也可以用：

# sheet = workbook.active
sheet = workbook["Sheet1"]

# 3.获取表格内容所在的范围
print(sheet.dimensions)

# 4.获取某个单元格的具体内容
# 这边提供两种方法，注意都需要以cell.value形式输出具体值
# 指定坐标
cell = sheet['A1']
print(cell.value)
# 指定行列
cell = sheet.cell(row = 3 , column = 2)
print(cell.value)

# 5.获取某个单元格的行、列、坐标
print(cell.row, cell.column , cell.coordinate)

# 6.获取多个格子的值
# 指定坐标范围
cells = sheet['A1:B5']
# 指定列的值
cells = sheet['A']
cells = sheet['A:C']
# 指定行的值
cells = sheet[5]
cells = sheet[5:6]
# 这里也有一个细节，Excel中每一列由字母确定，是字符型；每一行由一个数字确定，是整型。
# 当然，上面的三种方法都是获取一堆表格，现在要输出每一个表格的值就需要遍历：
# for cellchild in cells:
#     print(cellchild.value)

# 指定返回取值
# for row in sheet.iter_rows(min_row = 1, max_row = 3,min_col = 2, max_col = 5):
#     print(row)
#     for cell in row:
#         print(cell.value)

# 7.读取所有的行
# for row in sheet.rows:
#     print(row)

# 写入单元格
cell = sheet['A1']
cell.value = 'test'

# 写入一行或者多行数据
# data1 = ['小明', 25 , "五年级"]
# sheet.append(data1)
# data2 = ['小江', 25 , "六年级"]
# sheet.append(data2)
data = [
    ['小明', 25 , "五年级"],
    ['小江', 25 , "六年级"]
]
# for i in data:
#     sheet.append(i)

# 将公式写入单元格保存
sheet['K11'] = '=AVERAGE(K1:K10)' 

# 设置样式
# Font(name名称 , size大小, bold粗体, italic斜体, color颜色)
font = Font(name='微软雅黑', size=5,bold=True, italic=True, color='FF000000')
cell.font = font
# 设置对齐方式
# Alignment(horizontal=水平对齐, vertical=垂直对齐, text_rotation=字体倾斜度, wrap_text=自动换行)
# 水平对齐:distributed, justify, center, left, fill, centerContinuous, right, general
# 垂直对齐:bottom, distributed, justify, center, top
alignment = Alignment(horizontal='center', vertical='center', text_rotation=45, wrap_text=True)
cell.alignment = alignment
# 设置边框样式
# 边线样式:double, mediumDashDotDot, slantDashDot, dashDotDot, dotted, hair, mediumDashed, dashed, dashDot, thin, mediumDashDot, medium, thick
side = Side(style='thin', color='FF000000')
border = Border(left=side , right=side , top=side , bottom=side)
cell.border = border
#设置单元格填充格式
cell = sheet['A2']
pattern_fill = PatternFill(patternType='solid',fgColor='99ccff')
cell.fill = pattern_fill
cell2 = sheet['A3']
gradient_fill = GradientFill(stop=('FFFFFF','99ccff','000000'))
cell2.fill = gradient_fill
# 设置行高和列宽
sheet.row_dimensions[1].height = 50 
sheet.column_dimensions['C'].width = 20 

# 6. 单元格合并与取消

# 合并
sheet.merge_cells('A1:B2') 
sheet.merge_cells(start_row=1, start_column=3, 
                  end_row=2, end_column=4)

# 取消合并
sheet.unmerge_cells('A1:B2') 
sheet.unmerge_cells(start_row=1, start_column=3, 
                    end_row=2, end_column=4)

# 保存Excel
workbook.save(fileName)
